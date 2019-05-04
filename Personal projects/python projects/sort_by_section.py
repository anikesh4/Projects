import openpyxl

while True:
	def find_row(sheet,column,section_name):
	    rows_with_sections=[]
	    #print("\n\n\n\n")
	    #print(sheet[column])
	    #print("\n\n\n\n")
	    for cells in sheet[column]:
	        if cells.value==section_name:
	            rows_with_sections.append(cells.row)

	    return rows_with_sections

	def copy_paste_row(parent_sheet, parent_row, child_sheet, child_row, num_cols):
	    #num_cols=parent_sheet.max_column
	    #please pass the row number
	    row_data=['#']
	    for cell in range(1,num_cols):
	        row_data.append(parent_sheet.cell(parent_row, column = cell).value)
	      #  print(parent_sheet.cell(parent_row, column = cell).value)
	        
	    #print('\n\n############################################\n\n')
	    for cell in range(1,num_cols):
	        child_sheet.cell(child_row, column = cell).value = row_data[cell]
	       # print(child_sheet.cell(child_row, column = cell).value) 2019-04-02T1545_Grades-CHEM123.AllLabs
	       # CHEM 123 L07


	while True:
		try:
			workbook_name=input("Please enter the exact file name of the workbook (without the extention: ")
			wb= openpyxl.load_workbook(workbook_name + ".xlsx")
		except FileNotFoundError:
			print("File not found, please enter the correct name:  ")
			continue
		except:
			print("file could not be opened, make sure it is in the right location")
			continue
		else:
			print("workbook opened successfully !")
			break

	section_col=input("Please enter the exact column letter:  ")
	parent_sheet= wb.get_sheet_by_name(wb.get_sheet_names()[0])
	section= input('please enter the exact name of the section: ')
	new_sheet= wb.create_sheet(section)
	rows_needed=[]

	#print(f"\n\nsection_col: {type(section_col)}, parent_sheet: {type(parent_sheet)}, section:{type(section)}\n\n")
	rows_needed = find_row(parent_sheet,section_col,section)
	title_rows=[1,2,3]


	child_row= 1
	num_cols= parent_sheet.max_column

	for parent_rows in title_rows:
		copy_paste_row(parent_sheet, parent_rows, new_sheet, child_row, num_cols)
		child_row+=1 

	
	for parent_rows in rows_needed:
		copy_paste_row(parent_sheet, parent_rows, new_sheet, child_row, num_cols)
		child_row+=1

	print("A new sheet has been created with only section "+ section + "students in a new workbook")
	wb.save('pyhton_sorted_workbook.xlsx')


	again=" "
	while again!='yes' and again !='no':
		again= input("Do you have any more sections to sort, 'yes' or 'no'??:  ").lower()

	if again=='no':
		print("See you next time!!!")
		break